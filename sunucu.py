#!/usr/bin/env python3
"""
Firma Otomasyonu — backend (Flask'sız, sadece stdlib + scraping kodu)
---------------------------------------------------------------------
Çalıştır:
  /Users/emrearslan/firma-dogrulama/.venv/bin/python sunucu.py
  -> http://127.0.0.1:8300

Uç noktalar:
  GET  /                 -> index.html (senin arayüzün)
  POST /api/yukle        -> ham dosya gövdesi (X-Dosya-Adi başlığı) -> firma adları
  GET  /api/akis?token=  -> SSE: her firma için canlı sonuç
  GET  /api/indir?token= -> doldurulmuş .xlsx
"""
import io, os, csv, json, time, secrets, sys
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, unquote
from openpyxl import load_workbook, Workbook

# mevcut otomasyon mantığını içe aktar
sys.path.insert(0, "/Users/emrearslan/firma-dogrulama")
import firma_doldur as fd  # resmi_site_bul, iletisim_bul

BURADA = os.path.dirname(os.path.abspath(__file__))
CAP = 20          # canlı demo için satır sınırı (yavaşlamasın)
OTURUM = {}       # token -> {"firmalar": [...], "xlsx": bytes}
BASLIK = ("firma", "firma adı", "firma adi", "firma ismi", "company",
          "company name", "unvan", "ünvan", "isim", "ad")


def isimleri_cikar(dosya_adi: str, ham: bytes):
    ext = dosya_adi.lower().rsplit(".", 1)[-1]
    isimler = []
    if ext in ("xlsx", "xlsm", "xls"):
        wb = load_workbook(io.BytesIO(ham), read_only=True, data_only=True)
        for row in wb.active.iter_rows(min_row=1, max_col=1, values_only=True):
            if row[0] is not None and str(row[0]).strip():
                isimler.append(str(row[0]).strip())
    else:  # csv
        metin = ham.decode("utf-8-sig", errors="replace")
        for row in csv.reader(io.StringIO(metin)):
            if row and row[0].strip():
                isimler.append(row[0].strip())
    # başlık satırını ele
    if isimler and isimler[0].lower() in BASLIK:
        isimler = isimler[1:]
    return isimler


def xlsx_uret(sonuclar):
    wb = Workbook(); ws = wb.active; ws.title = "Sonuc"
    ws.append(["Firma Adı", "Web Sitesi", "E-posta", "Telefon", "Durum"])
    for r in sonuclar:
        ws.append([r["firma"], r["site"], r["eposta"], r["telefon"], r["durum"]])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *a, **k):
        super().__init__(*a, directory=BURADA, **k)

    def log_message(self, *a):  # sessiz
        pass

    # ---- POST: dosya yükleme ----
    def do_POST(self):
        if urlparse(self.path).path != "/api/yukle":
            self.send_error(404); return
        try:
            n = int(self.headers.get("Content-Length", 0))
            ham = self.rfile.read(n)
            ad = unquote(self.headers.get("X-Dosya-Adi", "liste.xlsx"))
            isimler = isimleri_cikar(ad, ham)
            token = secrets.token_hex(8)
            OTURUM[token] = {"firmalar": isimler}
            self._json({"token": token, "toplam": len(isimler),
                        "firmalar": isimler[:CAP], "sinir": CAP})
        except Exception as e:
            self._json({"hata": str(e)}, 400)

    # ---- GET: statik + SSE + indir ----
    def do_GET(self):
        yol = urlparse(self.path).path
        sorgu = parse_qs(urlparse(self.path).query)
        if yol == "/api/akis":
            return self._akis(sorgu.get("token", [""])[0])
        if yol == "/api/indir":
            return self._indir(sorgu.get("token", [""])[0])
        if yol in ("/", ""):
            self.path = "/index.html"
        return super().do_GET()

    def _akis(self, token):
        if token not in OTURUM:
            self.send_error(404); return
        self.send_response(200)
        self.send_header("Content-Type", "text/event-stream")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Connection", "keep-alive")
        self.end_headers()

        def emit(event, data):
            self.wfile.write(f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n".encode())
            self.wfile.flush()

        firmalar = OTURUM[token]["firmalar"][:CAP]
        sonuclar = []
        for i, firma in enumerate(firmalar):
            try:
                site = fd.resmi_site_bul(firma)
                eposta, telefon = fd.iletisim_bul(site, firma)
            except Exception:
                site = eposta = telefon = ""
            r = {"index": i, "firma": firma, "site": site,
                 "eposta": eposta, "telefon": telefon,
                 "durum": "Tamamlandı" if site else "Site bulunamadı"}
            sonuclar.append(r)
            emit("satir", r)
            time.sleep(0.3)
        OTURUM[token]["xlsx"] = xlsx_uret(sonuclar)
        emit("bitti", {"sayi": len(sonuclar)})

    def _indir(self, token):
        veri = OTURUM.get(token, {}).get("xlsx")
        if not veri:
            self.send_error(404); return
        self.send_response(200)
        self.send_header("Content-Type",
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", 'attachment; filename="firmalar_dolu.xlsx"')
        self.send_header("Content-Length", str(len(veri)))
        self.end_headers()
        self.wfile.write(veri)

    def _json(self, obj, code=200):
        veri = json.dumps(obj, ensure_ascii=False).encode()
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(veri)))
        self.end_headers()
        self.wfile.write(veri)


if __name__ == "__main__":
    print("→ http://127.0.0.1:8300   (Ctrl+C ile durdur)")
    ThreadingHTTPServer(("127.0.0.1", 8300), Handler).serve_forever()
